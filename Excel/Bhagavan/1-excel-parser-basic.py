from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, colors, PatternFill
import dump_cell

wb = load_workbook('revenue.xlsx')

#print worksheet names
print wb.get_sheet_names()
for wsheet in wb.get_sheet_names():
    print wsheet

#opening worksheet in workbook
wsheet = wb.get_sheet_by_name('sales')

#max rows and cols
rcount =  wsheet.max_row
ccount =  wsheet.max_column
print "max rows filled :", rcount
print "max cols filled :", ccount


#print cell value
print "A2 :", wsheet['A2'].value 
print "A3 :", wsheet['C3'].value 

#print in row and col format
print "wsheet[1][0] :", wsheet[1][0].value 
print "wsheet[2][1] :", wsheet[2][1].value 

#print rows and columns
for row in range(1, rcount):
    flag = 1
    for col in range (0, ccount):
        if (wsheet[row][col].value != None):
            print (str(wsheet[row][col].value)),
            flag = 0

    print ""
    if (flag == 1):
       break;

print ""
#print rows and columns
for row in range(1, rcount):
    if not any(cell.value for cell in wsheet[row]):
        print "The %d row is empty" % row
        break

    for col in range (0, ccount):
        if (wsheet[row][col].value != None):
            print (str(wsheet[row][col].value)),

    print ""

#Highlight cell in case cell value is Vinay
#Change cell font size to 12, color Red, bold
for row in range(1, rcount):
    if not any(cell.value for cell in wsheet[row]):
        print "The %d row is empty" % row
        break

    for col in range (0, ccount):
        if (wsheet[row][col].value != None and wsheet[row][col].value == "Vinay" ):
            print (wsheet[row][col].font)
            wsheet[row][col].font = Font(size=12, bold=True, color=colors.RED)
            print ""

#Highlight row in case column value is Ashish
#Change cell font size to 12, colore GREEN, bold
for row in range(1, rcount):
    if not any(cell.value for cell in wsheet[row]):
        print "The %d row is empty" % row
        break

    for col in range (0, ccount):
        if (wsheet[row][col].value != None and wsheet[row][col].value == "Ashish" ):
            for cell in wsheet[row:row]:
                cell.font = Font(size=12, bold=True, color=colors.GREEN)

#Highlight row back ground in case column value is Kavitha
#Change cell back ground to yellow
for row in range(1, rcount):
    if not any(cell.value for cell in wsheet[row]):
        print "The %d row is empty" % row
        break

    for col in range (0, ccount):
        if (wsheet[row][col].value != None and wsheet[row][col].value == "Kavitha" ):
            print (wsheet[row][col].fill)
            print (str(wsheet[row][col].value))
            for cell in wsheet[row:row]:
				cell.fill = PatternFill(fill_type="solid", start_color=colors.GREEN)

wb.save('revenue.xlsx')
exit(1)

'''
for word in dir(wsheet['C3']):
    print "print \"%-20s :\", cell.%s" %  (word, word)
'''

dump_cell.dump_cell_properties(wsheet[1][0])

wb.save('revenue.xlsx')
exit(1)
