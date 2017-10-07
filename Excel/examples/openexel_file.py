#http://nullege.com/codes/search/openpyxl.workbook.Workbook.get_sheet_by_name

import openpyxl

wb=openpyxl.load_workbook("CSVFile_11kb.xlsx")

#print(type(wb))
'''
#====================================
sheet_names=wb.get_sheet_names() # print sheet names

print(type(sheet_names))

for name in sheet_names:
     print(name)
#=====================================

#Ex:-1
#wb.get_sheet_by_name Exampls

sheet_names=wb.get_sheet_names() # print sheet names

print(type(sheet_names))

for sheet_name in sheet_names:
     name=wb.get_sheet_by_name(sheet_name)
     print("name",name)
     print(name.title)
#=======================================
'''



shetnames=wb.get_sheet_by_name("Sheet")

print(type(shetnames))

rows=shetnames.max_row
columns=shetnames.max_column
print(rows,columns)
'''
for c in range(1,rows):
    d=shetnames.cell(row=1,column=c)
    print(d.value)

print("\n\n")

for r in range(1,columns):
    d=shetnames.cell(row=r,column=3)
    print(d.value)
'''
for r in range(1,rows+1):
    for c in range (1,columns+1):
        d=shetnames.cell(row=r,column=c)
        print('%-30s'%d.value,end=' ')
    print(' ')